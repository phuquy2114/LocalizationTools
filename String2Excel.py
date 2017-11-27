"""checktext - Check that a text file has macintosh-style newlines"""

import sys
import os
import string
from openpyxl.styles import colors
from openpyxl.styles import Font, PatternFill, Color
from openpyxl.styles.borders import Border, Side
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.comments import Comment
import datetime
import xml.etree.ElementTree as ET
import localizable
import pygsheets

PLATFORM = sys.argv[1]
RES_PATH = sys.argv[2]
SPREAD_ID = sys.argv[3]
try:
    MERGE_OPTION = sys.argv[4]
except:
    MERGE_OPTION = "their"
try:
    SYN_SPREAD = sys.argv[4]
except:
    SYN_SPREAD = "their"

# Download WorkBook
# Parameter:
#   SheetID: the google sheet ID
#Return:
#   the name of SpreadSheet, if download success
#   None, if download fail
def DownloadWorkBook(SpreadID):
    result = None
    try:
        gc = pygsheets.authorize()
        sh = gc.open_by_key(SpreadID)
        content_format = pygsheets.ExportType.MS_Excel
        sh_name = sh.title + ".xlsx"
        sh_0 = sh.title + "0.xlsx"
        num_sheets = len(sh.worksheets())
        sh.export(fformat=content_format,filename=sh_name)
        print 'num sheets: ' + str(num_sheets)
        if os.path.exists(sh_name):
            os.remove(sh_name)
        if os.path.exists(sh_0):
            os.rename(sh_0,sh_name)
        for x in range(1, num_sheets):
            tmp = sh.title + str(x) + ".xlsx"
            if os.path.exists(tmp):
                os.remove(tmp)
        result = sh_name
    except:
        result = None
    return result

def getSheetName(filename, platform):
    sheetName = ""
    if platform == "android":
        sheetName = filename[6:len(filename)-4]
    else:
        sheetName = filename[0:len(filename)-8]
    return sheetName

def getStringPath(string_folder, filename, platform):
    strPath = ""
    if platform == "android":
        strPath = os.path.join(string_folder,"values")
        strPath = os.path.join(strPath, filename)
    else:
        strPath = os.path.join(string_folder,"Base.lproj")
        strPath = os.path.join(strPath, filename)
    return strPath

def findXMLValue(xml_tre,name,tag):
    for chil in xml_tre:
        if chil.attrib['name'] == name and chil.tag == tag:
            return chil.text
    return None
def xmlNode2String(xml_node):
    #str_key = chil.attrib['name']
    str_text = ""
    if not (xml_node.text is None):
        str_text = xml_node.text
        str_text.replace("\\'","\'")
    if len(xml_node.getchildren()) <= 0:
        return str_text
    #if xml_node.tag == "string":
        #print xml_node.attrib['name']
        #str_text = '<format>'
    #else:
    if xml_node.tag.find("-array") >= 0:
        str_text = '<' +xml_node.tag+ ' name="' + xml_node.attrib['name'] + '">'
        
    for chil in xml_node:
        #if xml_node.tag == "string-array":
            #str_text +='\n \t'
        str_text += "<" + chil.tag
        for attrName, attrValue in chil.attrib.items():
            str_text += ' ' + attrName +'="' + attrValue + '"'
        str_text += ">" + xmlNode2String(chil) + "</" + chil.tag + ">"
    #if xml_node.tag == "string":
        #str_text += '</format>'
    #else:
    if xml_node.tag.find("-array") >= 0:
        str_text += '</' +xml_node.tag+ '>'
    #print "note: " + str_text
    return str_text
    
def fillSheetAndroid(string_path, ws):
    #print "process: " + string_path
    #check relative file
    base_folder = os.path.dirname(string_path)
    base_folder = os.path.join(base_folder, os.pardir)
    file_name = os.path.basename(string_path)

    # Vietnamese
    vi_path = os.path.join(base_folder,"values-vi")
    vi_path = os.path.join(vi_path,file_name)
    vi_resources = None
    if os.path.exists(vi_path):
        vi_tree = ET.parse(vi_path)
        vi_resources = vi_tree.getroot()

    # Japanese
    ja_path = os.path.join(base_folder,"values-ja")
    ja_path = os.path.join(ja_path,file_name)
    ja_resources = None
    if os.path.exists(ja_path):
        ja_tree = ET.parse(ja_path)
        ja_resources = ja_tree.getroot()
    
    # Fill table data
    tree = ET.parse(string_path)
    resources = tree.getroot()
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    #resources = root.find('resources')
    if resources != None:
        row = 1
        for chil in resources:
            row +=1
            id_cell = "A" + str(row)
            description_cell = "B" + str(row)
            vi_cell = "D" + str(row)
            en_cell = "C" + str(row)
            ja_cell = "E" + str(row)
            str_key = chil.attrib['name']
            ws[id_cell] = str_key
            ws[id_cell].border = thin_border
            ws[id_cell].alignment = ws[id_cell].alignment.copy(wrapText=True)
            ws[description_cell].border = thin_border
            
            if chil.tag == "string":
                #ws[en_cell] = chil.text
                #string_value = chil.text;
                string_value = xmlNode2String(chil)
                #if string_value is None:
                    #string_value = "".join(chil.itertext())
                ws[en_cell] = string_value
                ws[en_cell].border = thin_border
                if vi_resources != None:
                    ws[vi_cell].value = findXMLValue(vi_resources,str_key,"string")
                ws[vi_cell].alignment = ws[vi_cell].alignment.copy(wrapText=True)
                ws[vi_cell].border = thin_border
                ws[vi_cell].alignment = ws[en_cell].alignment.copy(wrapText=True)
                if ja_resources != None:
                    ws[ja_cell].value = findXMLValue(ja_resources,str_key,"string")
                ws[ja_cell].border = thin_border
                ws[ja_cell].alignment = ws[ja_cell].alignment.copy(wrapText=True)
            else:
                #if chil.tag == "string-array":
                ws[description_cell].value = chil.tag
                ws[en_cell].value = xmlNode2String(chil)
    else:
        print 'resources is None'

def findValue(dic,key):
    for item in dic:
        if item['key'] == key:
            return item['value']
    return None
def fillSheetIOS(string_path, ws):
    #check relative file
    base_folder = os.path.dirname(string_path)
    base_folder = os.path.join(base_folder, os.pardir)
    file_name = os.path.basename(string_path)

    # Vietnamese
    vi_path = os.path.join(base_folder,"vi.lproj")
    vi_path = os.path.join(vi_path,file_name)
    vi_strings = None
    if os.path.exists(vi_path):
        vi_strings = localizable.parse_strings(filename=vi_path)

    # Japanese
    ja_path = os.path.join(base_folder,"ja.lproj")
    ja_path = os.path.join(ja_path,file_name)
    ja_strings = None
    if os.path.exists(ja_path):
        ja_strings = localizable.parse_strings(filename=ja_path)
        
    
    # Fill table data
    strings = localizable.parse_strings(filename=string_path)
    row = 1
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    haveError = False
    for item in strings:
        row +=1
        id_cell = "A" + str(row)
        description_cell = "B" + str(row)
        vi_cell = "D" + str(row)
        en_cell = "C" + str(row)
        ja_cell = "E" + str(row)
        str_key = item['key']
        str_comment = item['comment']
        if isinstance(str_key, unicode):
            print str_key + ':' +str_comment
            ws[description_cell].value = str_comment
        ws[id_cell].value = str_key
        ws[id_cell].border = thin_border
        ws[id_cell].alignment = ws[id_cell].alignment.copy(wrapText=True)
        ws[description_cell].border = thin_border
        ws[en_cell].value = item['value']
        ws[en_cell].border = thin_border
        ws[en_cell].alignment = ws[vi_cell].alignment.copy(wrapText=True)
        
        if vi_strings != None:
            try:
               ws[vi_cell].value = findValue(vi_strings,str_key)
            except:
                haveError = True
                redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                ws[vi_cell].fill = redFill
        ws[vi_cell].border = thin_border
        ws[vi_cell].alignment = ws[en_cell].alignment.copy(wrapText=True)

        if ja_strings != None:
            ws[ja_cell].value = findValue(ja_strings,str_key)
        ws[ja_cell].border = thin_border
        ws[ja_cell].alignment = ws[ja_cell].alignment.copy(wrapText=True)
    if haveError:
        ws.sheet_properties.tabColor = 'FF0000'

def addSheet(wb, string_folder, filename, platform):
    #reduce the sheetname length
    new_sheetname = getSheetName(filename,platform)
    
    string_path = getStringPath(string_folder,filename,platform)
    #print 'create sheet for: ' + xml_path
    ws = wb.create_sheet(new_sheetname)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    if platform == "android":
        fillSheetAndroid(string_path,ws)
    else:
        fillSheetIOS(string_path,ws)
        
    # Create table title
    ws['A1'] = "TEXT ID"
    ws['B1'] = "Description"
    ws['C1'] = "EN"
    ws['D1'] = "VI"
    ws['E1'] = "JA"
    headerFill = PatternFill(start_color='FFFF5500', end_color='FFFF5500', fill_type='solid')
    ws['A1'].fill = ws['B1'].fill = ws['C1'].fill = ws['D1'].fill = ws['E1'].fill = headerFill
    ws['A1'].border = ws['B1'].border = ws['C1'].border = ws['D1'].border = ws['E1'].border = thin_border
    ws.column_dimensions['A'].width = ws.column_dimensions['C'].width = ws.column_dimensions['D'].width = ws.column_dimensions['E'].width = 50

    

# Compare 2 cell
# return True if the cell are different
# merge_option: "their", "mine", None
def compare_cell(mine, their, merge, merge_option):
    # format cell
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    merge.border = thin_border
    merge.alignment = merge.alignment.copy(wrapText=True)
    
    if merge_option is None:
        merge.value = mine.value
        return False;
    if mine.value == their.value:
        merge.value = mine.value
        return False;

    #handle conflict
    if merge_option == "their":
        if their.value in [None,'']:
            merge.value = mine.value
            greenFill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')
            merge.fill = greenFill
        else:
            merge.value = their.value
            comment = Comment(mine.value, 'mine')
            merge.comment = comment
            pinkFill = PatternFill(start_color='FFFF00FF', end_color='FFFF00FF', fill_type='solid')
            merge.fill = pinkFill
    else:
        merge.value = mine.value
        comment = Comment(their.value, 'their')
        merge.comment = comment
        redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        merge.fill = redFill;
    return True;
    
def compare_sheet(mine, their, merge, merge_option):
    haveConflict = False
    # for the new sheet
    if their is None:
        for row in mine.iter_rows():
            for cell in row:
                cell_name = cell.column + str(cell.row)
                merge_cell = merge[cell_name]
                merge_cell.value = cell.value
        merge.sheet_properties.tabColor = '00FF00'
    else:
        for row in mine.iter_rows():
            for cell in row:
                cell_name = cell.column + str(cell.row)
                their_cell = their[cell_name]
                merge_cell = merge[cell_name]
                if (compare_cell(cell,their_cell,merge_cell,merge_option)):
                    haveConflict = True

    # format header
    headerFill = PatternFill(start_color='FFFF5500', end_color='FFFF5500', fill_type='solid')
    for col in merge.iter_cols(min_col=None, max_col=None, max_row=1):
        for cell in col:
            cell.fill = headerFill
    merge.column_dimensions['A'].width = merge.column_dimensions['C'].width = merge.column_dimensions['D'].width = merge.column_dimensions['E'].width = 50


    if haveConflict:
        #print "Have Conflict"
        if merge_option == "their":
            merge.sheet_properties.tabColor = 'FF00FF'
        else:
            merge.sheet_properties.tabColor = 'FF0000'

def compare_workbook(mine, their, merge, merge_option):
    for sheet in mine.worksheets:
        sheet_name = sheet.title
        print "comapare sheet: " + sheet_name
        merge_sheet = merge.create_sheet(sheet_name)
        their_sheet = None
        try:
            their_sheet = their[sheet_name]
        except:
            their_sheet = None
        #if not (their_sheet is None):
        compare_sheet(sheet, their_sheet, merge_sheet, merge_option);

def string2exel(platform, res_folder, exel_name):
    strings_folder = ""
    wb = Workbook()
    delete_sheet = wb.active
    wb.remove_sheet(delete_sheet)
    if platform == "android":
        strings_folder =os.path.join(res_folder,"values")
        for filename in os.listdir(strings_folder):
            if (filename.find("string") == 0) and (filename.rfind(".xml") == len(filename)-4):
                addSheet(wb,res_folder,filename,"android")
            else:
                continue
    else:
        strings_folder =os.path.join(res_folder,"Base.lproj")
        for filename in os.listdir(strings_folder):
            if filename.rfind(".strings") == len(filename)-8:
                addSheet(wb,res_folder,filename,"ios")
            else:
                continue
    delete_sheet = None
    wb.save(exel_name)
    wb = None

#def

#synchronyze/upload the wookbook to GoogleSheet
#Paragrams:
#   wb: the workbook yo want to upload
#   googleSheetID: The ID of Google Spread
def synWorkBook2Spread(wb, googleSpreadID):
    gc = pygsheets.authorize()
    # Open spreadsheet and then workseet
    sh = gc.open_by_key(googleSpreadID)
    dumy_sheet = None
    dumy_name = 'dumy_sheet_____'
    try :
        dumy_sheet = sh.worksheet_by_title(dumy_name)
    except:
        sh.add_worksheet(dumy_name, rows=1, cols=1)
    for ws in sh.worksheets():
        if ws != dumy_sheet:
            sh.del_worksheet(ws)
    
    for sheet in wb.worksheets:
        sheet_name = sheet.title
        num_row = sheet.max_row
        num_col = sheet.max_column
        print "Syn Sheet: " + sheet_name + "num_col: " + str(num_col)
        wks = sh.add_worksheet(sheet_name, rows=num_row, cols=num_col)
        
        #sh.batch_start()
        #for i in range(1, num_row):
            #for j in range(1, num_col):
               #wks.cell((i,j)).value = sheet.cell(row=j, column=i).value
        #sh.batch_stop()
            
        for i in range(1, num_col+1):
            #print sheet.cell(row=1, column=i).value
            cell_values = []
            for j in range(1, num_row+1):
                cell_values.append(sheet.cell(row=j, column=i).value)
            wks.update_col(i,cell_values)
        wks.adjust_column_width(0,end=5,pixel_size=300)

        # format header
        print "Syn color..."
        for i in range(1, num_col+1):
            for j in range(1, num_row+1):
                wb_cell = sheet.cell(row=j, column=i)
                fill_color = wb_cell.fill.bgColor.rgb
                if len(fill_color) < 8:
                    fill_color = "00" + fill_color
                cell_color = tuple(round(float(int(fill_color[i:i+2], 16))/255,1) for i in (2,4,6))
                #print fill_color + ":" + str(cell_color)
                if cell_color != (1.0, 1.0, 1.0) and cell_color != (0.0, 0.0, 0.0):
                    s_cell = wks.cell((j, i))
                    s_cell.color = cell_color
                    wb_comment = wb_cell.comment
                    if not (wb_comment is None):
                        s_cell.note = wb_comment.text
        wb_sheet_color = merge.sheet_properties.tabColor
    if not (dumy_sheet is None):
        sh.del_worksheet(dumy_sheet)
                   
def mergeStringSpread(platform, res_folder, googleSpreadID, merge_option, syn2Spread):
    exel_name = DownloadWorkBook(googleSpreadID)
    if exel_name != None:
        name_split = exel_name.split(".")
        mine_exel_name = name_split[0] + ".mine." + name_split[1]
        their_excel_name = name_split[0] + ".their." + name_split[1]
        merge_exel_name = name_split[0] + ".merge." + name_split[1]

        # clearn
        if os.path.exists(their_excel_name):
            os.remove(their_excel_name)
        if os.path.exists(mine_exel_name):
            os.remove(mine_exel_name)
        if os.path.exists(merge_exel_name):
            os.remove(merge_exel_name)

        # export String to excel
        string2exel(platform, res_folder, mine_exel_name)

        #merge:
        their_wb = load_workbook(exel_name)
        #their_wb.active = 1
        #their_wb.save(their_excel_name)
        mine_wb = load_workbook(mine_exel_name)
        merge_wb = Workbook()
        delete_sheet = merge_wb.active
        merge_wb.remove_sheet(delete_sheet)
        delete_sheet = None

        compare_workbook(mine_wb, their_wb, merge_wb, merge_option)            
        merge_wb.save(merge_exel_name)
        if syn2Spread == True:
            synWorkBook2Spread(merge_wb,googleSpreadID)
        #merge_wb = None

if __name__ == '__main__':
    # main()
    mergeStringSpread(PLATFORM,RES_PATH,SPREAD_ID,MERGE_OPTION,SYN_SPREAD)
    sys.exit(0)
