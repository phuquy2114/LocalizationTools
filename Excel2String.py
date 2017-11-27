"""checktext - Check that a text file has macintosh-style newlines"""

import sys
import os
import string
from openpyxl.styles import colors
from openpyxl.styles import Font, PatternFill, Color
from openpyxl.styles.borders import Border, Side
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.comments import Comment
import datetime
import xml.etree.ElementTree as ET
import localizable
import HTMLParser
import codecs
import io
import re

PLATFORM = sys.argv[1]
EXCEL_FILE = sys.argv[2]
RES_PATH = sys.argv[3]

narrow = len('\U00010000') == 2
def ord_skip(s, ndx):
    ch = ord(s[ndx])
    if ch >= 0xd800 and ch <= 0xdbff:
        hi = (ch - 0xd800) << 10
        ch2 = ord(s[ndx + 1])
        lo = (ch2 - 0xdc00)
        return (0x10000 + (hi | lo), 2)
    return (ch, 1)
_c_escapes = {
    0x07: 'a',
    0x08: 'b',
    0x0c: 'f',
    0x0a: 'n',
    0x0d: 'r',
    0x09: 't',
    0x0b: 'v',
    0x22: '"',
    0x5c: '\\'
}
    
if narrow:
    _esc_re = re.compile('([\x00-\x1f\x7f-\x9f\u200e\u200f\u2028-\u202e\ufe00-\ufe0f"\\\\]|\udb40[\udd00-\uddef])')
else:
    _esc_re = re.compile('[\x00-\x1f\x7f-\x9f\u200e\u200f\u2028-\u202e\ufe00-\ufe0f\U000e0100-\U000e01ef"\\\\]')

def escape_string(s):
    result = []
    pos = 0
    end = len(s)
    while pos < end:
        m = _esc_re.search(s, pos)
        if m:
            result.append(s[pos:m.start(0)])
            ch = m.group(0)
            cp = ord_skip(ch, 0)[0]
            pos = m.end(0)

            if cp in _c_escapes:
                result.append('\\%s' % _c_escapes[cp])
            elif cp <= 0xff:
                result.append('\\x%02x' % cp)
            elif cp <= 0xffff:
                result.append('\\u%04x' % cp)
            else:
                result.append('\\U%08x' % cp)
        else:
            result.append(s[pos:])
            pos = end
    return ''.join(result)

class LocalizedString(object):
    __slots__ = ['key', 'str_id', 'en', 'vi', 'ja', 'comment']
    
    def __init__(self, key, str_id, en, vi=None, ja=None, comment=None):
        # The string to translate
        self.key = key
        self.str_id = str_id

        # English
        self.en = en

        # Vietnamese
        self.vi = vi

        # Japanese
        self.ja = ja

        # The comment, if any
        self.comment = comment

    def __eq__(self, other):
        return self.key == other.key and self.str_id == other.str_id and self.en == other.en and self.vi == other.vi and self.ja == other.ja and self.comment == other.comment

    def __ne__(self, other):
        return self.key != other.key or self.str_id != other.str_id or self.en != other.en or self.vi != other.vi or self.ja != other.ja or self.comment != other.comment
        
    def __repr__(self):
        return '%r' % self.en

class StringTable(object):
    def __init__(self):
        self.strings = {}

    def __eq__(self, other):
        return self.strings == other.strings

    def __ne__(self, other):
        return self.strings != other.strings
        
    def __getitem__(self, key):
        return self.strings[key]

    def __setitem__(self, key, str_id, en, vi=None, ja=None, comment=None):
        self.store(LocalizedString(key, str_id, en, vi, ja, comment))

    def __repr__(self):
        return '%r' % self.strings

    def lookup(self, key):
        return self.strings.get(key, None)

    def store(self, localized_string):
        cur = self.strings.get(localized_string.key, None)
        if cur:
            if localized_string.comment:
                if cur.comment:
                    cur.comment += '\n' + localized_string.comment
                else:
                    cur.comment = localized_string.comment
            cur.str_id = localized_string.str_id
            cur.en = localized_string.en
            cur.vi = localized_string.vi
            cur.ja = localized_string.ja
            #print "duplicate:" + localized_string.str_id
        else:
            self.strings[localized_string.key] = localized_string
            #print "add:" + localized_string.str_id

    # If called as StringTable.read(), will construct a new object and read
    # the strings into that.  Otherwise reads into the stringtable "self".
    #alsoconstruct
    def read(self, sheet):
        self.strings = {}
        max_row = sheet.max_row
        max_columm = sheet.max_column
        id_columm = 1
        ds_colume = 2
        en_columm = 3
        vi_columm = 4
        ja_columm = 5
        #find columm
        for i in range(1,max_columm+1):
            if sheet.cell(row=1, column=i).value.lower() == 'text id':
                id_columm = i
            if sheet.cell(row=1, column=i).value.lower() == 'en':
                en_columm = i
            if sheet.cell(row=1, column=i).value.lower() == 'vi':
                vi_columm = i
            if sheet.cell(row=1, column=i).value.lower() == 'ja':
                ja_columm = i
            if sheet.cell(row=1, column=i).value.lower() == 'description':
                ds_colume = i
        count = 0
        for row in range(2,max_row+1):
            str_id = sheet.cell(row=row, column=id_columm).value
            str_description = sheet.cell(row=row, column=ds_colume).value
            str_en = sheet.cell(row=row, column=en_columm).value
            #str_en = str_en.repalce('\n', '\\n')
            str_en= str_en.replace(u'\n', '\\n')
            str_en= str_en.replace(u'\t', '\\t')
            str_vi = sheet.cell(row=row, column=vi_columm).value
            if not (str_vi is None):
                str_vi= str_vi.replace(u'\n', '\\n')
                str_vi= str_vi.replace(u'\t', '\\t')
            str_ja = sheet.cell(row=row, column=ja_columm).value
            if not (str_ja is None):
                str_ja= str_ja.replace(u'\n', '\\n')
                str_ja= str_ja.replace(u'\t', '\\t')
            key = '{:5d}'.format(count)
            
            self.store(LocalizedString(key,str_id, str_en, str_vi, str_ja, str_description))
            count +=1

    def write(self, out_path, file_name, encoding='utf_8'):
        #check and create folders
        en_strings_path = os.path.join(out_path,"Base.lproj")
        vi_strings_path = os.path.join(out_path,"vi.lproj")
        ja_strings_path = os.path.join(out_path,"ja.lproj")
        if not os.path.exists(en_strings_path):
            os.mkdir(en_strings_path)
        if not os.path.exists(vi_strings_path):
            os.mkdir(vi_strings_path)
        if not os.path.exists(ja_strings_path):
            os.mkdir(ja_strings_path)
        en_file_name = os.path.join(en_strings_path, file_name)
        vi_file_name = os.path.join(vi_strings_path, file_name)
        ja_file_name = os.path.join(ja_strings_path, file_name)
        
        if isinstance(en_file_name, basestring):
            en_file_name = io.open(en_file_name, 'wb')
        if isinstance(vi_file_name, basestring):
            vi_file_name = io.open(vi_file_name, 'wb')
        if isinstance(ja_file_name, basestring):
            ja_file_name = io.open(ja_file_name, 'wb')

        en_writer_factory = codecs.getwriter(encoding)
        vi_writer_factory = codecs.getwriter(encoding)
        ja_writer_factory = codecs.getwriter(encoding)

        en_writer = en_writer_factory(en_file_name)
        vi_writer = vi_writer_factory(vi_file_name)
        ja_writer = ja_writer_factory(ja_file_name)

        if encoding != 'utf_8':
            writer.write('\ufeff')
        
        keys = self.strings.keys()
        keys.sort()

        first = True
        for k in keys:
            #print "key" + k
            if first:
                first = False
            else:
                en_writer.write('\n')
                vi_writer.write('\n')
                ja_writer.write('\n')
            ls = self.strings[k]
            #print ls.str_id
            if ls.comment:
                en_writer.write('/* %s */\n' % ls.comment)
                vi_writer.write('/* %s */\n' % ls.comment)
                ja_writer.write('/* %s */\n' % ls.comment)
            #else:
                #writer.write('/* No description */\n')
            en_writer.write('"%s" = "%s";\n' % (ls.str_id,ls.en))
            vi_writer.write('"%s" = "%s";\n' % (ls.str_id,ls.vi))
            ja_writer.write('"%s" = "%s";\n' % (ls.str_id,ls.ja))

def Colume2String(sheet, id_columm, desciption_columm, lange_columm, out_file):
    max_row = sheet.max_row +1
    resources = ET.Element('resources')
    resources.text = '\n\t'
    tree = ET.ElementTree(element=resources)
    last_element = None
    for row in range(2,max_row):
        str_id = sheet.cell(row=row, column=id_columm).value
        element_text = sheet.cell(row=row, column=lange_columm).value
        str_tag = 'string'
        element_tree = None
        if element_text != None:
            if isinstance(element_text, bool):
                element_text = str(element_text)
            if isinstance(element_text, long):
                element_text = str(element_text)
            
            if isinstance(element_text, str):
                element_text = element_text.replace("\'","\\'")
            if (not isinstance(element_text, unicode)) and (not isinstance(element_text, bool)) and (not isinstance(element_text, long)):
                print element_text
                element_text = unicode(element_text, 'utf-8', errors='replace')

            try:
                #print element_text
                element_tree = ET.fromstring(element_text.encode('utf-8'))
            except UnicodeEncodeError as err:
                print 'UnicodeEncodeError: ' + element_text, err
            except TypeError:
                if element_text != None:
                    print 'TypeError' + element_text
            #except ET.ParseError:
                #str_error = 'ParseError: ' + element_text
            except:
                #print element_text
                element_tree = None
            
        #if sheet.cell(row=row, column=desciption_columm).value == 'string-array' and element_text != None:
        if element_tree != None:
            #str_tag = 'string-array'
            str_tag = element_tree.tag
            #print str_tag
            #print element_text
            #parser = ET.XMLParser()
            isArray = False
            if str_tag.find("-array") >= 0:
                isArray = True
            #array_node = ET.fromstring(element_text)
            if isArray == True:
                mum_elements = len(element_tree.getchildren())
                element_tree.text = '\n\t\t'
                count = 0
                for chil in element_tree:
                    count += 1
                    if count < mum_elements:
                        chil.tail = '\n\t\t'
                    else:
                        chil.tail = '\n\t'
                element_tree.tail = '\n\t'
                resources.append(element_tree)
                last_element = element_tree
            else:
                str_tag = 'string'
                string_element = ET.SubElement(resources, str_tag, name=str_id)
                string_element.append(element_tree)
                string_element.tail = '\n\t'
                last_element = string_element
        else:
            #print element_text
            string_element = ET.SubElement(resources, str_tag, name=str_id)
            string_element.text = element_text
            string_element.tail = '\n\t'
            last_element = string_element
    if last_element != None:
        last_element.tail = '\n'
    tree.write(out_file,encoding="utf-8",xml_declaration=True)
def Sheet2IosString(sheet, out_path):
    file_name = sheet.title + ".strings"
    string_table = StringTable()
    string_table.read(sheet)
    string_table.write(out_path, file_name)

def Sheet2AndroidString(sheet, out_path):
    #check and create folders
    en_xml_path = os.path.join(out_path,"values")
    vi_xml_path = os.path.join(out_path,"values-vi")
    ja_xml_path = os.path.join(out_path,"values-ja")
    if not os.path.exists(en_xml_path):
        os.mkdir(en_xml_path)
    if not os.path.exists(vi_xml_path):
        os.mkdir(vi_xml_path)
    if not os.path.exists(ja_xml_path):
        os.mkdir(ja_xml_path)
    
    id_columm = 1
    ds_colume = 2
    en_columm = 3
    vi_columm = 4
    ja_columm = 5
    #find columm
    for i in range(1,sheet.max_column+1):
        if sheet.cell(row=1, column=i).value.lower() == 'text id':
            id_columm = i
        if sheet.cell(row=1, column=i).value.lower() == 'en':
            en_columm = i
        if sheet.cell(row=1, column=i).value.lower() == 'vi':
            vi_columm = i
        if sheet.cell(row=1, column=i).value.lower() == 'ja':
            ja_columm = i
        if sheet.cell(row=1, column=i).value.lower() == 'description':
            ds_colume = i
        
    file_name = "string" + sheet.title + ".xml"

    #english
    en_xml_file = os.path.join(en_xml_path,file_name)
    Colume2String(sheet,id_columm,ds_colume,en_columm,en_xml_file)
    #Vietnamese
    vi_xml_file = os.path.join(vi_xml_path,file_name)
    Colume2String(sheet,id_columm,ds_colume,vi_columm,vi_xml_file)
    #Japanese
    ja_xml_file = os.path.join(ja_xml_path,file_name)
    Colume2String(sheet,id_columm,ds_colume,ja_columm,ja_xml_file)

def Excel2String(platform, excel_file, out_path):
    wb = load_workbook(excel_file)
    for sheet in wb.worksheets:
        if platform == "android":
            Sheet2AndroidString(sheet, out_path)
        else:
            Sheet2IosString(sheet, out_path)

if __name__ == '__main__' :
    # main()
    Excel2String(PLATFORM,EXCEL_FILE,RES_PATH)
    sys.exit(0)
